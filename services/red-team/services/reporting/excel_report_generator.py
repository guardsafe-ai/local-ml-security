"""
Excel Report Generator for Enterprise Reporting
Generates comprehensive Excel reports for executives and auditors
"""

import logging
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import io
import base64

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """
    Generates comprehensive Excel reports for enterprise stakeholders
    """
    
    def __init__(self):
        """Initialize Excel report generator"""
        self.workbook = None
        self._setup_styles()
        logger.info("✅ Excel Report Generator initialized")
    
    def _setup_styles(self):
        """Setup cell styles for formatting"""
        try:
            # Header styles
            self.header_font = Font(bold=True, color="FFFFFF", size=12)
            self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            self.header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Subheader styles
            self.subheader_font = Font(bold=True, color="2F5597", size=11)
            self.subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            
            # Data styles
            self.data_font = Font(size=10)
            self.data_alignment = Alignment(horizontal="left", vertical="center")
            
            # Risk level styles
            self.critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            self.high_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            self.medium_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            self.low_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            
            # Border styles
            self.thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
        except Exception as e:
            logger.error(f"Style setup failed: {e}")
    
    def generate_executive_dashboard(self, 
                                   security_data: Dict[str, Any],
                                   output_path: str) -> bool:
        """
        Generate executive dashboard Excel report
        
        Args:
            security_data: Security assessment data
            output_path: Path to save Excel file
            
        Returns:
            True if successful
        """
        try:
            self.workbook = Workbook()
            
            # Remove default sheet
            self.workbook.remove(self.workbook.active)
            
            # Create sheets
            self._create_executive_summary_sheet(security_data)
            self._create_risk_overview_sheet(security_data)
            self._create_attack_analysis_sheet(security_data)
            self._create_compliance_status_sheet(security_data)
            self._create_recommendations_sheet(security_data)
            self._create_detailed_findings_sheet(security_data)
            
            # Save workbook
            self.workbook.save(output_path)
            
            logger.info(f"Executive dashboard generated: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Executive dashboard generation failed: {e}")
            return False
    
    def generate_audit_workbook(self, 
                              audit_data: Dict[str, Any],
                              output_path: str) -> bool:
        """
        Generate comprehensive audit workbook
        
        Args:
            audit_data: Audit trail and compliance data
            output_path: Path to save Excel file
            
        Returns:
            True if successful
        """
        try:
            self.workbook = Workbook()
            
            # Remove default sheet
            self.workbook.remove(self.workbook.active)
            
            # Create sheets
            self._create_audit_summary_sheet(audit_data)
            self._create_compliance_matrix_sheet(audit_data)
            self._create_evidence_tracking_sheet(audit_data)
            self._create_remediation_tracking_sheet(audit_data)
            self._create_control_testing_sheet(audit_data)
            self._create_findings_details_sheet(audit_data)
            
            # Save workbook
            self.workbook.save(output_path)
            
            logger.info(f"Audit workbook generated: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Audit workbook generation failed: {e}")
            return False
    
    def generate_technical_analysis_workbook(self, 
                                           technical_data: Dict[str, Any],
                                           output_path: str) -> bool:
        """
        Generate technical analysis workbook
        
        Args:
            technical_data: Technical analysis data
            output_path: Path to save Excel file
            
        Returns:
            True if successful
        """
        try:
            self.workbook = Workbook()
            
            # Remove default sheet
            self.workbook.remove(self.workbook.active)
            
            # Create sheets
            self._create_technical_summary_sheet(technical_data)
            self._create_attack_techniques_sheet(technical_data)
            self._create_model_performance_sheet(technical_data)
            self._create_vulnerability_details_sheet(technical_data)
            self._create_exploitability_analysis_sheet(technical_data)
            self._create_metrics_analysis_sheet(technical_data)
            
            # Save workbook
            self.workbook.save(output_path)
            
            logger.info(f"Technical analysis workbook generated: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Technical analysis workbook generation failed: {e}")
            return False
    
    def _create_executive_summary_sheet(self, data: Dict[str, Any]):
        """Create executive summary sheet"""
        try:
            ws = self.workbook.create_sheet("Executive Summary")
            
            # Title
            ws['A1'] = "ML Security Assessment - Executive Summary"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Report metadata
            row = 3
            metadata = [
                ["Report Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Assessment Period:", data.get('assessment_period', 'N/A')],
                ["Models Tested:", str(data.get('models_tested', 0))],
                ["Total Attacks:", str(data.get('total_attacks', 0))],
                ["Vulnerabilities Found:", str(data.get('vulnerabilities_found', 0))],
                ["Overall Risk Level:", data.get('overall_risk_level', 'Unknown')]
            ]
            
            for item in metadata:
                ws[f'A{row}'] = item[0]
                ws[f'B{row}'] = item[1]
                ws[f'A{row}'].font = self.subheader_font
                ws[f'B{row}'].font = self.data_font
                row += 1
            
            # Key findings
            row += 2
            ws[f'A{row}'] = "Key Findings"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            row += 1
            
            key_findings = data.get('key_findings', [])
            for i, finding in enumerate(key_findings[:10], 1):
                ws[f'A{row}'] = f"{i}. {finding}"
                ws[f'A{row}'].font = self.data_font
                row += 1
            
            # Risk summary
            row += 2
            ws[f'A{row}'] = "Risk Summary"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            row += 1
            
            risk_summary = data.get('risk_summary', {})
            risk_data = [
                ["Risk Level", "Count", "Percentage"],
                ["Critical", risk_summary.get('critical', 0), f"{risk_summary.get('critical_pct', 0)}%"],
                ["High", risk_summary.get('high', 0), f"{risk_summary.get('high_pct', 0)}%"],
                ["Medium", risk_summary.get('medium', 0), f"{risk_summary.get('medium_pct', 0)}%"],
                ["Low", risk_summary.get('low', 0), f"{risk_summary.get('low_pct', 0)}%"]
            ]
            
            for i, row_data in enumerate(risk_data):
                for j, cell_data in enumerate(row_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    if i == 0:  # Header row
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.header_alignment
                    else:
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                        # Color code risk levels
                        if j == 0 and i > 0:
                            risk_level = row_data[0].lower()
                            if risk_level == 'critical':
                                cell.fill = self.critical_fill
                            elif risk_level == 'high':
                                cell.fill = self.high_fill
                            elif risk_level == 'medium':
                                cell.fill = self.medium_fill
                            elif risk_level == 'low':
                                cell.fill = self.low_fill
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Executive summary sheet creation failed: {e}")
    
    def _create_risk_overview_sheet(self, data: Dict[str, Any]):
        """Create risk overview sheet"""
        try:
            ws = self.workbook.create_sheet("Risk Overview")
            
            # Title
            ws['A1'] = "Risk Overview & Analysis"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Risk distribution chart data
            row = 3
            risk_distribution = data.get('risk_distribution', {})
            
            # Create risk distribution table
            risk_data = [
                ["Risk Level", "Count", "Percentage", "Trend", "Status"],
                ["Critical", risk_distribution.get('critical', 0), f"{risk_distribution.get('critical_pct', 0)}%", 
                 risk_distribution.get('critical_trend', '→'), "🔴"],
                ["High", risk_distribution.get('high', 0), f"{risk_distribution.get('high_pct', 0)}%",
                 risk_distribution.get('high_trend', '→'), "🟠"],
                ["Medium", risk_distribution.get('medium', 0), f"{risk_distribution.get('medium_pct', 0)}%",
                 risk_distribution.get('medium_trend', '→'), "🟡"],
                ["Low", risk_distribution.get('low', 0), f"{risk_distribution.get('low_pct', 0)}%",
                 risk_distribution.get('low_trend', '→'), "🟢"]
            ]
            
            for i, row_data in enumerate(risk_data):
                for j, cell_data in enumerate(row_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    if i == 0:  # Header row
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.header_alignment
                    else:
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                        # Color code risk levels
                        if j == 0 and i > 0:
                            risk_level = row_data[0].lower()
                            if risk_level == 'critical':
                                cell.fill = self.critical_fill
                            elif risk_level == 'high':
                                cell.fill = self.high_fill
                            elif risk_level == 'medium':
                                cell.fill = self.medium_fill
                            elif risk_level == 'low':
                                cell.fill = self.low_fill
                row += 1
            
            # Add risk distribution chart
            chart_row = row + 2
            chart = BarChart()
            chart.title = "Risk Distribution"
            chart.x_axis.title = "Risk Level"
            chart.y_axis.title = "Count"
            
            data_range = Reference(ws, min_col=2, min_row=2, max_row=5, max_col=2)
            categories = Reference(ws, min_col=1, min_row=2, max_row=5)
            chart.add_data(data_range, titles_from_data=False)
            chart.set_categories(categories)
            
            ws.add_chart(chart, f'A{chart_row}')
            
            # Top vulnerabilities
            row = chart_row + 15
            ws[f'A{row}'] = "Top Vulnerabilities"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            row += 1
            
            top_vulnerabilities = data.get('top_vulnerabilities', [])
            vuln_headers = ["Rank", "Vulnerability", "Severity", "CVSS Score", "Status"]
            for j, header in enumerate(vuln_headers):
                cell = ws.cell(row=row, column=j+1, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            row += 1
            
            for i, vuln in enumerate(top_vulnerabilities[:10], 1):
                vuln_data = [
                    i,
                    vuln.get('name', 'Unknown'),
                    vuln.get('severity', 'Unknown'),
                    vuln.get('cvss_score', 'N/A'),
                    vuln.get('status', 'Open')
                ]
                for j, cell_data in enumerate(vuln_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    cell.font = self.data_font
                    cell.alignment = self.data_alignment
                    # Color code severity
                    if j == 2:  # Severity column
                        severity = str(cell_data).lower()
                        if 'critical' in severity:
                            cell.fill = self.critical_fill
                        elif 'high' in severity:
                            cell.fill = self.high_fill
                        elif 'medium' in severity:
                            cell.fill = self.medium_fill
                        elif 'low' in severity:
                            cell.fill = self.low_fill
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Risk overview sheet creation failed: {e}")
    
    def _create_attack_analysis_sheet(self, data: Dict[str, Any]):
        """Create attack analysis sheet"""
        try:
            ws = self.workbook.create_sheet("Attack Analysis")
            
            # Title
            ws['A1'] = "Attack Analysis & Statistics"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Attack statistics
            row = 3
            attack_stats = data.get('attack_statistics', {})
            
            stats_data = [
                ["Metric", "Value", "Previous Period", "Change", "Trend"],
                ["Total Attacks", attack_stats.get('total_attacks', 0), 
                 attack_stats.get('prev_total_attacks', 0),
                 attack_stats.get('attacks_change', 0), attack_stats.get('attacks_trend', '→')],
                ["Successful Attacks", attack_stats.get('successful_attacks', 0),
                 attack_stats.get('prev_successful_attacks', 0),
                 attack_stats.get('success_change', 0), attack_stats.get('success_trend', '→')],
                ["Success Rate", f"{attack_stats.get('success_rate', 0):.1f}%",
                 f"{attack_stats.get('prev_success_rate', 0):.1f}%",
                 f"{attack_stats.get('success_rate_change', 0):.1f}%", attack_stats.get('success_rate_trend', '→')],
                ["Avg Response Time", f"{attack_stats.get('avg_response_time', 0):.2f}ms",
                 f"{attack_stats.get('prev_avg_response_time', 0):.2f}ms",
                 f"{attack_stats.get('response_time_change', 0):.2f}ms", attack_stats.get('response_time_trend', '→')],
                ["Detection Rate", f"{attack_stats.get('detection_rate', 0):.1f}%",
                 f"{attack_stats.get('prev_detection_rate', 0):.1f}%",
                 f"{attack_stats.get('detection_rate_change', 0):.1f}%", attack_stats.get('detection_rate_trend', '→')]
            ]
            
            for i, row_data in enumerate(stats_data):
                for j, cell_data in enumerate(row_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    if i == 0:  # Header row
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.header_alignment
                    else:
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                row += 1
            
            # Attack categories breakdown
            row += 2
            ws[f'A{row}'] = "Attack Categories Breakdown"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            row += 1
            
            attack_categories = data.get('attack_categories', {})
            cat_headers = ["Category", "Count", "Percentage", "Success Rate", "Avg Severity"]
            for j, header in enumerate(cat_headers):
                cell = ws.cell(row=row, column=j+1, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            row += 1
            
            for category, stats in attack_categories.items():
                cat_data = [
                    category,
                    stats.get('count', 0),
                    f"{stats.get('percentage', 0):.1f}%",
                    f"{stats.get('success_rate', 0):.1f}%",
                    f"{stats.get('avg_severity', 0):.2f}"
                ]
                for j, cell_data in enumerate(cat_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    cell.font = self.data_font
                    cell.alignment = self.data_alignment
                row += 1
            
            # Add attack categories pie chart
            chart_row = row + 2
            pie_chart = PieChart()
            pie_chart.title = "Attack Categories Distribution"
            
            data_range = Reference(ws, min_col=2, min_row=row-len(attack_categories), max_row=row-1, max_col=2)
            categories = Reference(ws, min_col=1, min_row=row-len(attack_categories), max_row=row-1)
            pie_chart.add_data(data_range, titles_from_data=False)
            pie_chart.set_categories(categories)
            
            ws.add_chart(pie_chart, f'A{chart_row}')
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Attack analysis sheet creation failed: {e}")
    
    def _create_compliance_status_sheet(self, data: Dict[str, Any]):
        """Create compliance status sheet"""
        try:
            ws = self.workbook.create_sheet("Compliance Status")
            
            # Title
            ws['A1'] = "Compliance Status & Framework Analysis"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Compliance frameworks
            row = 3
            compliance = data.get('compliance_status', {})
            
            compliance_data = [
                ["Framework", "Status", "Score", "Last Assessment", "Next Due", "Gap Count"],
                ["SOC 2", compliance.get('soc2', {}).get('status', 'N/A'),
                 f"{compliance.get('soc2', {}).get('score', 0)}%",
                 compliance.get('soc2', {}).get('last_assessment', 'N/A'),
                 compliance.get('soc2', {}).get('next_due', 'N/A'),
                 compliance.get('soc2', {}).get('gap_count', 0)],
                ["ISO 27001", compliance.get('iso27001', {}).get('status', 'N/A'),
                 f"{compliance.get('iso27001', {}).get('score', 0)}%",
                 compliance.get('iso27001', {}).get('last_assessment', 'N/A'),
                 compliance.get('iso27001', {}).get('next_due', 'N/A'),
                 compliance.get('iso27001', {}).get('gap_count', 0)],
                ["OWASP LLM Top 10", compliance.get('owasp', {}).get('status', 'N/A'),
                 f"{compliance.get('owasp', {}).get('score', 0)}%",
                 compliance.get('owasp', {}).get('last_assessment', 'N/A'),
                 compliance.get('owasp', {}).get('next_due', 'N/A'),
                 compliance.get('owasp', {}).get('gap_count', 0)]
            ]
            
            for i, row_data in enumerate(compliance_data):
                for j, cell_data in enumerate(row_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    if i == 0:  # Header row
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.header_alignment
                    else:
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                        # Color code status
                        if j == 1:  # Status column
                            status = str(cell_data).lower()
                            if 'compliant' in status or 'passed' in status:
                                cell.fill = self.low_fill
                            elif 'partial' in status or 'in progress' in status:
                                cell.fill = self.medium_fill
                            elif 'non-compliant' in status or 'failed' in status:
                                cell.fill = self.critical_fill
                row += 1
            
            # Compliance recommendations
            row += 2
            ws[f'A{row}'] = "Compliance Recommendations"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            row += 1
            
            recommendations = data.get('compliance_recommendations', [])
            for i, rec in enumerate(recommendations[:15], 1):
                ws[f'A{row}'] = f"{i}. {rec}"
                ws[f'A{row}'].font = self.data_font
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Compliance status sheet creation failed: {e}")
    
    def _create_recommendations_sheet(self, data: Dict[str, Any]):
        """Create recommendations sheet"""
        try:
            ws = self.workbook.create_sheet("Recommendations")
            
            # Title
            ws['A1'] = "Security Recommendations & Action Plan"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Immediate actions
            row = 3
            ws[f'A{row}'] = "Immediate Actions (0-30 days)"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            row += 1
            
            immediate_headers = ["Priority", "Action", "Owner", "Due Date", "Status", "Notes"]
            for j, header in enumerate(immediate_headers):
                cell = ws.cell(row=row, column=j+1, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            row += 1
            
            immediate_actions = data.get('immediate_recommendations', [])
            for i, action in enumerate(immediate_actions[:10], 1):
                action_data = [
                    "High" if i <= 3 else "Medium",
                    action,
                    "TBD",
                    (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"),
                    "Pending",
                    ""
                ]
                for j, cell_data in enumerate(action_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    cell.font = self.data_font
                    cell.alignment = self.data_alignment
                    # Color code priority
                    if j == 0:  # Priority column
                        if cell_data == "High":
                            cell.fill = self.critical_fill
                        elif cell_data == "Medium":
                            cell.fill = self.medium_fill
                row += 1
            
            # Short-term recommendations
            row += 2
            ws[f'A{row}'] = "Short-term Actions (1-3 months)"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            row += 1
            
            for j, header in enumerate(immediate_headers):
                cell = ws.cell(row=row, column=j+1, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            row += 1
            
            short_term_actions = data.get('short_term_recommendations', [])
            for i, action in enumerate(short_term_actions[:10], 1):
                action_data = [
                    "Medium" if i <= 5 else "Low",
                    action,
                    "TBD",
                    (datetime.now() + timedelta(days=90)).strftime("%Y-%m-%d"),
                    "Pending",
                    ""
                ]
                for j, cell_data in enumerate(action_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    cell.font = self.data_font
                    cell.alignment = self.data_alignment
                    # Color code priority
                    if j == 0:  # Priority column
                        if cell_data == "Medium":
                            cell.fill = self.medium_fill
                        elif cell_data == "Low":
                            cell.fill = self.low_fill
                row += 1
            
            # Long-term recommendations
            row += 2
            ws[f'A{row}'] = "Long-term Actions (3-12 months)"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            row += 1
            
            for j, header in enumerate(immediate_headers):
                cell = ws.cell(row=row, column=j+1, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            row += 1
            
            long_term_actions = data.get('long_term_recommendations', [])
            for i, action in enumerate(long_term_actions[:10], 1):
                action_data = [
                    "Low",
                    action,
                    "TBD",
                    (datetime.now() + timedelta(days=365)).strftime("%Y-%m-%d"),
                    "Pending",
                    ""
                ]
                for j, cell_data in enumerate(action_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    cell.font = self.data_font
                    cell.alignment = self.data_alignment
                    # Color code priority
                    if j == 0:  # Priority column
                        cell.fill = self.low_fill
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Recommendations sheet creation failed: {e}")
    
    def _create_detailed_findings_sheet(self, data: Dict[str, Any]):
        """Create detailed findings sheet"""
        try:
            ws = self.workbook.create_sheet("Detailed Findings")
            
            # Title
            ws['A1'] = "Detailed Security Findings"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Findings table
            row = 3
            findings = data.get('detailed_findings', [])
            
            if findings:
                findings_headers = ["ID", "Severity", "Category", "Description", "CVSS Score", "Status", "Remediation"]
                for j, header in enumerate(findings_headers):
                    cell = ws.cell(row=row, column=j+1, value=header)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                row += 1
                
                for finding in findings[:50]:  # Limit to 50 findings
                    finding_data = [
                        finding.get('id', 'N/A'),
                        finding.get('severity', 'N/A'),
                        finding.get('category', 'N/A'),
                        finding.get('description', 'N/A')[:100] + '...' if len(finding.get('description', '')) > 100 else finding.get('description', 'N/A'),
                        finding.get('cvss_score', 'N/A'),
                        finding.get('status', 'Open'),
                        finding.get('remediation', 'TBD')
                    ]
                    for j, cell_data in enumerate(finding_data):
                        cell = ws.cell(row=row, column=j+1, value=cell_data)
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                        # Color code severity
                        if j == 1:  # Severity column
                            severity = str(cell_data).lower()
                            if 'critical' in severity:
                                cell.fill = self.critical_fill
                            elif 'high' in severity:
                                cell.fill = self.high_fill
                            elif 'medium' in severity:
                                cell.fill = self.medium_fill
                            elif 'low' in severity:
                                cell.fill = self.low_fill
                    row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Detailed findings sheet creation failed: {e}")
    
    def _create_audit_summary_sheet(self, data: Dict[str, Any]):
        """Create audit summary sheet"""
        try:
            ws = self.workbook.create_sheet("Audit Summary")
            
            # Title
            ws['A1'] = "ML Security Audit Summary"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Audit metadata
            row = 3
            audit_metadata = [
                ["Audit Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Audit Period:", data.get('audit_period', 'N/A')],
                ["Auditor:", data.get('auditor', 'ML Security Platform')],
                ["Audit Scope:", data.get('audit_scope', 'Full ML Security Assessment')],
                ["Compliance Frameworks:", data.get('compliance_frameworks', 'SOC 2, ISO 27001')],
                ["Total Controls Tested:", str(data.get('controls_tested', 0))],
                ["Controls Passed:", str(data.get('controls_passed', 0))],
                ["Controls Failed:", str(data.get('controls_failed', 0))],
                ["Overall Compliance Score:", f"{data.get('compliance_score', 0)}%"]
            ]
            
            for item in audit_metadata:
                ws[f'A{row}'] = item[0]
                ws[f'B{row}'] = item[1]
                ws[f'A{row}'].font = self.subheader_font
                ws[f'B{row}'].font = self.data_font
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Audit summary sheet creation failed: {e}")
    
    def _create_compliance_matrix_sheet(self, data: Dict[str, Any]):
        """Create compliance matrix sheet"""
        try:
            ws = self.workbook.create_sheet("Compliance Matrix")
            
            # Title
            ws['A1'] = "Compliance Framework Matrix"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Compliance matrix
            row = 3
            compliance_matrix = data.get('compliance_matrix', {})
            
            if compliance_matrix:
                matrix_headers = ["Control", "SOC 2", "ISO 27001", "OWASP", "Status", "Evidence"]
                for j, header in enumerate(matrix_headers):
                    cell = ws.cell(row=row, column=j+1, value=header)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                row += 1
                
                for control, frameworks in compliance_matrix.items():
                    matrix_data = [
                        control,
                        "✓" if frameworks.get('soc2', False) else "✗",
                        "✓" if frameworks.get('iso27001', False) else "✗",
                        "✓" if frameworks.get('owasp', False) else "✗",
                        "PASS" if all(frameworks.values()) else "FAIL",
                        frameworks.get('evidence', 'N/A')
                    ]
                    for j, cell_data in enumerate(matrix_data):
                        cell = ws.cell(row=row, column=j+1, value=cell_data)
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                        # Color code status
                        if j == 4:  # Status column
                            if cell_data == "PASS":
                                cell.fill = self.low_fill
                            else:
                                cell.fill = self.critical_fill
                    row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Compliance matrix sheet creation failed: {e}")
    
    def _create_evidence_tracking_sheet(self, data: Dict[str, Any]):
        """Create evidence tracking sheet"""
        try:
            ws = self.workbook.create_sheet("Evidence Tracking")
            
            # Title
            ws['A1'] = "Audit Evidence & Documentation"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Evidence summary
            row = 3
            evidence = data.get('evidence', {})
            
            evidence_data = [
                ["Evidence Type", "Count", "Location", "Status", "Last Updated"],
                ["Audit Logs", evidence.get('audit_logs', 0), evidence.get('audit_logs_path', 'N/A'), "Complete", datetime.now().strftime("%Y-%m-%d")],
                ["Test Results", evidence.get('test_results', 0), evidence.get('test_results_path', 'N/A'), "Complete", datetime.now().strftime("%Y-%m-%d")],
                ["Screenshots", evidence.get('screenshots', 0), evidence.get('screenshots_path', 'N/A'), "Complete", datetime.now().strftime("%Y-%m-%d")],
                ["Config Files", evidence.get('config_files', 0), evidence.get('config_files_path', 'N/A'), "Complete", datetime.now().strftime("%Y-%m-%d")],
                ["Network Traces", evidence.get('network_traces', 0), evidence.get('network_traces_path', 'N/A'), "Complete", datetime.now().strftime("%Y-%m-%d")]
            ]
            
            for i, row_data in enumerate(evidence_data):
                for j, cell_data in enumerate(row_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    if i == 0:  # Header row
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.header_alignment
                    else:
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Evidence tracking sheet creation failed: {e}")
    
    def _create_remediation_tracking_sheet(self, data: Dict[str, Any]):
        """Create remediation tracking sheet"""
        try:
            ws = self.workbook.create_sheet("Remediation Tracking")
            
            # Title
            ws['A1'] = "Remediation Tracking & Status"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Remediation tracking
            row = 3
            remediation = data.get('remediation_tracking', {})
            
            rem_headers = ["Finding ID", "Priority", "Assigned To", "Due Date", "Status", "Progress", "Notes"]
            for j, header in enumerate(rem_headers):
                cell = ws.cell(row=row, column=j+1, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            row += 1
            
            remediation_items = remediation.get('items', [])
            for item in remediation_items[:30]:  # Limit to 30 items
                rem_data = [
                    item.get('finding_id', 'N/A'),
                    item.get('priority', 'N/A'),
                    item.get('assigned_to', 'TBD'),
                    item.get('due_date', 'TBD'),
                    item.get('status', 'Open'),
                    f"{item.get('progress', 0)}%",
                    item.get('notes', '')
                ]
                for j, cell_data in enumerate(rem_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    cell.font = self.data_font
                    cell.alignment = self.data_alignment
                    # Color code priority and status
                    if j == 1:  # Priority column
                        priority = str(cell_data).lower()
                        if 'critical' in priority or 'high' in priority:
                            cell.fill = self.critical_fill
                        elif 'medium' in priority:
                            cell.fill = self.medium_fill
                        elif 'low' in priority:
                            cell.fill = self.low_fill
                    elif j == 4:  # Status column
                        status = str(cell_data).lower()
                        if 'closed' in status or 'resolved' in status:
                            cell.fill = self.low_fill
                        elif 'in progress' in status:
                            cell.fill = self.medium_fill
                        elif 'open' in status or 'pending' in status:
                            cell.fill = self.critical_fill
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Remediation tracking sheet creation failed: {e}")
    
    def _create_control_testing_sheet(self, data: Dict[str, Any]):
        """Create control testing sheet"""
        try:
            ws = self.workbook.create_sheet("Control Testing")
            
            # Title
            ws['A1'] = "Security Control Testing Results"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Control testing results
            row = 3
            control_testing = data.get('control_testing', {})
            
            control_headers = ["Control ID", "Control Name", "Test Method", "Result", "Evidence", "Tester", "Date"]
            for j, header in enumerate(control_headers):
                cell = ws.cell(row=row, column=j+1, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
            row += 1
            
            control_results = control_testing.get('results', [])
            for result in control_results[:50]:  # Limit to 50 results
                control_data = [
                    result.get('control_id', 'N/A'),
                    result.get('control_name', 'N/A'),
                    result.get('test_method', 'N/A'),
                    result.get('result', 'N/A'),
                    result.get('evidence', 'N/A'),
                    result.get('tester', 'N/A'),
                    result.get('date', 'N/A')
                ]
                for j, cell_data in enumerate(control_data):
                    cell = ws.cell(row=row, column=j+1, value=cell_data)
                    cell.font = self.data_font
                    cell.alignment = self.data_alignment
                    # Color code result
                    if j == 3:  # Result column
                        result_val = str(cell_data).lower()
                        if 'pass' in result_val or 'compliant' in result_val:
                            cell.fill = self.low_fill
                        elif 'fail' in result_val or 'non-compliant' in result_val:
                            cell.fill = self.critical_fill
                        elif 'partial' in result_val:
                            cell.fill = self.medium_fill
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Control testing sheet creation failed: {e}")
    
    def _create_findings_details_sheet(self, data: Dict[str, Any]):
        """Create findings details sheet"""
        try:
            ws = self.workbook.create_sheet("Findings Details")
            
            # Title
            ws['A1'] = "Detailed Security Findings"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Findings details
            row = 3
            findings = data.get('detailed_findings', [])
            
            if findings:
                findings_headers = ["ID", "Severity", "Category", "Description", "CVSS Score", "Status", "Remediation", "Evidence"]
                for j, header in enumerate(findings_headers):
                    cell = ws.cell(row=row, column=j+1, value=header)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                row += 1
                
                for finding in findings[:100]:  # Limit to 100 findings
                    finding_data = [
                        finding.get('id', 'N/A'),
                        finding.get('severity', 'N/A'),
                        finding.get('category', 'N/A'),
                        finding.get('description', 'N/A'),
                        finding.get('cvss_score', 'N/A'),
                        finding.get('status', 'Open'),
                        finding.get('remediation', 'TBD'),
                        finding.get('evidence', 'N/A')
                    ]
                    for j, cell_data in enumerate(finding_data):
                        cell = ws.cell(row=row, column=j+1, value=cell_data)
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                        # Color code severity
                        if j == 1:  # Severity column
                            severity = str(cell_data).lower()
                            if 'critical' in severity:
                                cell.fill = self.critical_fill
                            elif 'high' in severity:
                                cell.fill = self.high_fill
                            elif 'medium' in severity:
                                cell.fill = self.medium_fill
                            elif 'low' in severity:
                                cell.fill = self.low_fill
                    row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Findings details sheet creation failed: {e}")
    
    def _create_technical_summary_sheet(self, data: Dict[str, Any]):
        """Create technical summary sheet"""
        try:
            ws = self.workbook.create_sheet("Technical Summary")
            
            # Title
            ws['A1'] = "Technical Security Analysis Summary"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Technical metadata
            row = 3
            tech_metadata = [
                ["Report Type:", "Technical Deep Dive Analysis"],
                ["Analysis Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Analysis Period:", data.get('analysis_period', 'N/A')],
                ["Attack Techniques:", str(data.get('attack_techniques', 0))],
                ["Models Analyzed:", str(data.get('models_analyzed', 0))],
                ["Vulnerabilities Found:", str(data.get('vulnerabilities_found', 0))],
                ["Exploits Generated:", str(data.get('exploits_generated', 0))],
                ["Technical Risk Score:", f"{data.get('technical_risk_score', 0)}/10"]
            ]
            
            for item in tech_metadata:
                ws[f'A{row}'] = item[0]
                ws[f'B{row}'] = item[1]
                ws[f'A{row}'].font = self.subheader_font
                ws[f'B{row}'].font = self.data_font
                row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Technical summary sheet creation failed: {e}")
    
    def _create_attack_techniques_sheet(self, data: Dict[str, Any]):
        """Create attack techniques sheet"""
        try:
            ws = self.workbook.create_sheet("Attack Techniques")
            
            # Title
            ws['A1'] = "Attack Techniques Analysis"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Attack techniques data
            row = 3
            techniques = data.get('attack_techniques', {})
            
            if techniques:
                tech_headers = ["Technique", "Success Rate", "Avg Time", "Complexity", "Stealth", "Detection Rate"]
                for j, header in enumerate(tech_headers):
                    cell = ws.cell(row=row, column=j+1, value=header)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                row += 1
                
                for technique, stats in techniques.items():
                    tech_data = [
                        technique,
                        f"{stats.get('success_rate', 0):.1f}%",
                        f"{stats.get('avg_time', 0):.2f}s",
                        stats.get('complexity', 'N/A'),
                        stats.get('stealth', 'N/A'),
                        f"{stats.get('detection_rate', 0):.1f}%"
                    ]
                    for j, cell_data in enumerate(tech_data):
                        cell = ws.cell(row=row, column=j+1, value=cell_data)
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                    row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Attack techniques sheet creation failed: {e}")
    
    def _create_model_performance_sheet(self, data: Dict[str, Any]):
        """Create model performance sheet"""
        try:
            ws = self.workbook.create_sheet("Model Performance")
            
            # Title
            ws['A1'] = "Model Performance Analysis"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Model performance data
            row = 3
            performance = data.get('model_performance', {})
            
            if performance:
                perf_headers = ["Model", "Accuracy", "Precision", "Recall", "F1-Score", "Vulnerability Level", "Risk Score"]
                for j, header in enumerate(perf_headers):
                    cell = ws.cell(row=row, column=j+1, value=header)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                row += 1
                
                for model, stats in performance.items():
                    perf_data = [
                        model,
                        f"{stats.get('accuracy', 0):.3f}",
                        f"{stats.get('precision', 0):.3f}",
                        f"{stats.get('recall', 0):.3f}",
                        f"{stats.get('f1_score', 0):.3f}",
                        stats.get('vulnerability_level', 'N/A'),
                        f"{stats.get('risk_score', 0):.2f}"
                    ]
                    for j, cell_data in enumerate(perf_data):
                        cell = ws.cell(row=row, column=j+1, value=cell_data)
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                        # Color code vulnerability level
                        if j == 5:  # Vulnerability level column
                            vuln_level = str(cell_data).lower()
                            if 'critical' in vuln_level or 'high' in vuln_level:
                                cell.fill = self.critical_fill
                            elif 'medium' in vuln_level:
                                cell.fill = self.medium_fill
                            elif 'low' in vuln_level:
                                cell.fill = self.low_fill
                    row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Model performance sheet creation failed: {e}")
    
    def _create_vulnerability_details_sheet(self, data: Dict[str, Any]):
        """Create vulnerability details sheet"""
        try:
            ws = self.workbook.create_sheet("Vulnerability Details")
            
            # Title
            ws['A1'] = "Detailed Vulnerability Analysis"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Vulnerability details
            row = 3
            vulnerabilities = data.get('vulnerability_details', [])
            
            if vulnerabilities:
                vuln_headers = ["ID", "Name", "CVSS Score", "Severity", "Category", "Description", "Exploitability", "Remediation"]
                for j, header in enumerate(vuln_headers):
                    cell = ws.cell(row=row, column=j+1, value=header)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                row += 1
                
                for vuln in vulnerabilities[:50]:  # Limit to 50 vulnerabilities
                    vuln_data = [
                        vuln.get('id', 'N/A'),
                        vuln.get('name', 'Unknown'),
                        vuln.get('cvss_score', 'N/A'),
                        vuln.get('severity', 'N/A'),
                        vuln.get('category', 'N/A'),
                        vuln.get('description', 'N/A')[:100] + '...' if len(vuln.get('description', '')) > 100 else vuln.get('description', 'N/A'),
                        vuln.get('exploitability', 'N/A'),
                        vuln.get('remediation', 'N/A')
                    ]
                    for j, cell_data in enumerate(vuln_data):
                        cell = ws.cell(row=row, column=j+1, value=cell_data)
                        cell.font = self.data_font
                        cell.alignment = self.data_alignment
                        # Color code severity
                        if j == 3:  # Severity column
                            severity = str(cell_data).lower()
                            if 'critical' in severity:
                                cell.fill = self.critical_fill
                            elif 'high' in severity:
                                cell.fill = self.high_fill
                            elif 'medium' in severity:
                                cell.fill = self.medium_fill
                            elif 'low' in severity:
                                cell.fill = self.low_fill
                    row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Vulnerability details sheet creation failed: {e}")
    
    def _create_exploitability_analysis_sheet(self, data: Dict[str, Any]):
        """Create exploitability analysis sheet"""
        try:
            ws = self.workbook.create_sheet("Exploitability Analysis")
            
            # Title
            ws['A1'] = "Exploitability Assessment"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Exploitability data
            row = 3
            exploitability = data.get('exploitability_assessment', {})
            
            if exploitability:
                exp_data = [
                    ["Factor", "Score", "Description", "Impact"],
                    ["Overall Score", f"{exploitability.get('overall_score', 'N/A')}/10", "Overall exploitability rating", "High" if exploitability.get('overall_score', 0) > 7 else "Medium" if exploitability.get('overall_score', 0) > 4 else "Low"],
                    ["Attack Complexity", exploitability.get('attack_complexity', 'N/A'), "Complexity of executing the attack", "High" if exploitability.get('attack_complexity', 0) > 7 else "Medium" if exploitability.get('attack_complexity', 0) > 4 else "Low"],
                    ["Privileges Required", exploitability.get('privileges_required', 'N/A'), "Level of privileges needed", "High" if exploitability.get('privileges_required', 0) > 7 else "Medium" if exploitability.get('privileges_required', 0) > 4 else "Low"],
                    ["User Interaction", exploitability.get('user_interaction', 'N/A'), "Level of user interaction required", "High" if exploitability.get('user_interaction', 0) > 7 else "Medium" if exploitability.get('user_interaction', 0) > 4 else "Low"],
                    ["Scope", exploitability.get('scope', 'N/A'), "Scope of the vulnerability", "High" if exploitability.get('scope', 0) > 7 else "Medium" if exploitability.get('scope', 0) > 4 else "Low"]
                ]
                
                for i, row_data in enumerate(exp_data):
                    for j, cell_data in enumerate(row_data):
                        cell = ws.cell(row=row, column=j+1, value=cell_data)
                        if i == 0:  # Header row
                            cell.font = self.header_font
                            cell.fill = self.header_fill
                            cell.alignment = self.header_alignment
                        else:
                            cell.font = self.data_font
                            cell.alignment = self.data_alignment
                            # Color code impact
                            if j == 3:  # Impact column
                                impact = str(cell_data).lower()
                                if 'high' in impact:
                                    cell.fill = self.critical_fill
                                elif 'medium' in impact:
                                    cell.fill = self.medium_fill
                                elif 'low' in impact:
                                    cell.fill = self.low_fill
                    row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Exploitability analysis sheet creation failed: {e}")
    
    def _create_metrics_analysis_sheet(self, data: Dict[str, Any]):
        """Create metrics analysis sheet"""
        try:
            ws = self.workbook.create_sheet("Metrics Analysis")
            
            # Title
            ws['A1'] = "Security Metrics & KPIs"
            ws['A1'].font = Font(bold=True, size=16, color="2F5597")
            ws.merge_cells('A1:F1')
            
            # Metrics data
            row = 3
            metrics = data.get('security_metrics', {})
            
            if metrics:
                metrics_data = [
                    ["Metric", "Current Value", "Previous Value", "Change", "Target", "Status"],
                    ["Detection Rate", f"{metrics.get('detection_rate', 0):.1f}%", f"{metrics.get('prev_detection_rate', 0):.1f}%", f"{metrics.get('detection_rate_change', 0):.1f}%", "95%", "Good" if metrics.get('detection_rate', 0) >= 95 else "Needs Improvement"],
                    ["False Positive Rate", f"{metrics.get('false_positive_rate', 0):.1f}%", f"{metrics.get('prev_false_positive_rate', 0):.1f}%", f"{metrics.get('false_positive_rate_change', 0):.1f}%", "<5%", "Good" if metrics.get('false_positive_rate', 0) < 5 else "Needs Improvement"],
                    ["Response Time", f"{metrics.get('response_time', 0):.2f}ms", f"{metrics.get('prev_response_time', 0):.2f}ms", f"{metrics.get('response_time_change', 0):.2f}ms", "<100ms", "Good" if metrics.get('response_time', 0) < 100 else "Needs Improvement"],
                    ["Coverage", f"{metrics.get('coverage', 0):.1f}%", f"{metrics.get('prev_coverage', 0):.1f}%", f"{metrics.get('coverage_change', 0):.1f}%", "100%", "Good" if metrics.get('coverage', 0) >= 100 else "Needs Improvement"]
                ]
                
                for i, row_data in enumerate(metrics_data):
                    for j, cell_data in enumerate(row_data):
                        cell = ws.cell(row=row, column=j+1, value=cell_data)
                        if i == 0:  # Header row
                            cell.font = self.header_font
                            cell.fill = self.header_fill
                            cell.alignment = self.header_alignment
                        else:
                            cell.font = self.data_font
                            cell.alignment = self.data_alignment
                            # Color code status
                            if j == 5:  # Status column
                                status = str(cell_data).lower()
                                if 'good' in status:
                                    cell.fill = self.low_fill
                                elif 'needs improvement' in status:
                                    cell.fill = self.medium_fill
                    row += 1
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Metrics analysis sheet creation failed: {e}")
    
    def generate_report_summary(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Generate a summary of Excel report generation capabilities
        
        Args:
            data: Report data
            
        Returns:
            Summary dictionary
        """
        try:
            summary = {
                'report_types': [
                    'Executive Dashboard',
                    'Audit Workbook',
                    'Technical Analysis Workbook'
                ],
                'features': [
                    'Multi-sheet Excel workbooks',
                    'Interactive charts and graphs',
                    'Color-coded risk levels',
                    'Conditional formatting',
                    'Auto-fitted columns',
                    'Professional styling',
                    'Comprehensive data tables',
                    'Trend analysis',
                    'Status tracking',
                    'Evidence management'
                ],
                'data_sources': [
                    'Security assessment results',
                    'Attack analysis data',
                    'Compliance status',
                    'Vulnerability details',
                    'Audit logs and evidence',
                    'Model performance metrics',
                    'Remediation tracking',
                    'Control testing results'
                ],
                'output_formats': ['Excel (.xlsx)'],
                'target_audiences': [
                    'C-Level Executives',
                    'Security Auditors',
                    'Compliance Officers',
                    'Technical Teams',
                    'Board Members',
                    'Risk Managers'
                ]
            }
            
            return summary
            
        except Exception as e:
            logger.error(f"Excel report summary generation failed: {e}")
            return {"error": str(e)}
